import os
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

THESIS_DIR  = r"C:\Users\BornLoser\Desktop\Assignment\Thesis"
STEP8_DIR   = os.path.join(THESIS_DIR, "04_Mutation_Processing", "Output", "Step8_Merging")
STEP2_DIR   = os.path.join(THESIS_DIR, "04_Mutation_Processing", "Output", "Step2_RefCheck")

IN_PATH       = os.path.join(STEP8_DIR, "lovd_flat_with_patients.tsv")
OFFSET_PATH   = os.path.join(STEP2_DIR, "lrg_offset_results.csv")
OUT_PATH      = os.path.join(STEP8_DIR, "lovd_review.xlsx")

OUT_OF_SCOPE = {"BTK", "SH2"}
PROTEIN_ONLY = {"BLNK", "CD79A", "RASGRP2", "STAT2"}

VARIANT_COLS = [
    "_gene", "_dedup_key", "_patient_count", "_variant_class",
    "chromosome", "VariantOnGenome/DNA",
    "transcriptid", "VariantOnTranscript/DNA",
    "VariantOnTranscript/RNA", "VariantOnTranscript/Protein",
    "_mane_select", "_source_track",
]

INDIVIDUAL_COLS = [
    "Individual/ID", "_gene", "_accession",
    "Individual/Gender", "_gender_note",
    "Individual/Age", "Individual/Origin/Geographic", "Individual/Remarks",
    "Phenotype/Disease", "Phenotype/Additional",
    "Reference/PubMed_ID", "Reference/Authors", "Reference/Title", "Reference/Location",
    "_allele_count",
]

HEADER_FILL  = PatternFill("solid", start_color="305496")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT   = Font(name="Calibri", bold=True, size=16, color="1F3864")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="1F3864")
WARN_FILL    = PatternFill("solid", start_color="FFF2CC")
THIN_BORDER  = Border(*([Side(style="thin", color="BFBFBF")] * 4))


def agg_unique_join(s):
    vals = sorted({v.strip() for v in s if isinstance(v, str) and v.strip()})
    return " | ".join(vals)


def classify_gender(g):
    if not g:
        return ""
    if " | " in g:
        return f"conflict: {g}"
    if g in {"M", "F"}:
        return ""
    return f"ambiguous: {g}"


def style_header(ws, row=1, ncols=None):
    if ncols is None:
        ncols = ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def autosize(ws, max_width=60):
    for c in range(1, ws.max_column + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def write_df(ws, df):
    ws.append(list(df.columns))
    style_header(ws, ncols=len(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([("" if pd.isna(v) else v) for v in row])
    ws.freeze_panes = "A2"
    autosize(ws)


def compute_excluded_genes(df_screening):
    if not os.path.exists(OFFSET_PATH):
        return [], [], []
    offset_df = pd.read_csv(OFFSET_PATH, dtype=str).fillna("")
    included = set(df_screening["_gene"].str.upper().unique())
    in_offset = set(offset_df["gene"].str.upper().unique())
    missing = in_offset - included

    excluded_oos     = sorted(missing & OUT_OF_SCOPE)
    excluded_protein = sorted(missing & PROTEIN_ONLY)
    excluded_nofasta = sorted(g for g in missing
                              if g not in OUT_OF_SCOPE and g not in PROTEIN_ONLY)
    return excluded_oos, excluded_protein, excluded_nofasta


def build_summary_sheet(ws, df_screening, df_variants, df_individuals,
                        fill_rates, track_counts, class_counts,
                        excluded_oos, excluded_protein, excluded_nofasta):
    ws["A1"] = "LOVD Migration Review — IDBases to LOVD 3.0"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Generated: {date.today().isoformat()}"
    ws["A2"].font = Font(italic=True, color="595959")

    r = 4
    ws.cell(row=r, column=1, value="Counts").font = SECTION_FONT
    r += 1
    counts = [
        ("Unique genes included",        df_variants["_gene"].nunique()),
        ("Unique variants (gene+dedup_key)", len(df_variants)),
        ("Unique individuals",           len(df_individuals)),
        ("Patient-variant cases (rows)", len(df_screening)),
    ]
    for label, val in counts:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value=(
        "Note: 'Patient-variant cases' includes every distinct (patient, variant) pair. "
        "When the same variant appears in multiple individuals, each individual is counted separately. "
        "The internal dedup step only collapsed cross-track redundancy (same patient-variant resolved by both NM_MANE and NG_IDRefseq Mutalyzer tracks)."
    ))
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 45
    r += 2

    ws.cell(row=r, column=1, value="Gene exclusions").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="Category")
    ws.cell(row=r, column=2, value="Count")
    ws.cell(row=r, column=3, value="Genes")
    style_header(ws, row=r, ncols=3)
    r += 1
    for label, genes in [
        ("Intentionally out of scope (handled separately)", excluded_oos),
        ("Protein-only source data (no genomic/transcript variants)", excluded_protein),
        ("Reference FASTA not confirmed in Step 2 (no_fasta)", excluded_nofasta),
    ]:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=len(genes))
        ws.cell(row=r, column=3, value=", ".join(genes) if genes else "—")
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Variant class distribution").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="Class")
    ws.cell(row=r, column=2, value="Variants")
    style_header(ws, row=r, ncols=2)
    r += 1
    for klass, n in class_counts.items():
        ws.cell(row=r, column=1, value=klass)
        ws.cell(row=r, column=2, value=n)
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Source track distribution").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="Track")
    ws.cell(row=r, column=2, value="Variants")
    style_header(ws, row=r, ncols=2)
    r += 1
    for track, n in track_counts.items():
        ws.cell(row=r, column=1, value=track)
        ws.cell(row=r, column=2, value=n)
        r += 1
    r += 1

    ws.cell(row=r, column=1, value=f"Fill rates (n={len(df_screening)} patient-variant rows)").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="LOVD column")
    ws.cell(row=r, column=2, value="Filled")
    ws.cell(row=r, column=3, value="%")
    style_header(ws, row=r, ncols=3)
    r += 1
    for col, (filled, total) in fill_rates.items():
        ws.cell(row=r, column=1, value=col)
        ws.cell(row=r, column=2, value=f"{filled}/{total}")
        ws.cell(row=r, column=3, value=round(100 * filled / total, 1) if total else 0)
        if total and (filled / total) < 0.30:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = WARN_FILL
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Pipeline provenance").font = SECTION_FONT
    r += 1
    notes = [
        "Source: 131 IDBases immunodeficiency gene databases (BMC Lund University)",
        "Pipeline: 7 steps — extraction, ref confirmation, BLAST, BED, LiftOver (GRCh38), Mutalyzer 3, merge",
        "Mutalyzer: 3 tracks (NM_MANE / NG_IDRefseq / NM_IDRefseq), priority merge",
        "Reference build: GRCh38 (hg38)",
        "MANE Select: applied where NM_MANE track succeeded",
        "Deduplication: only cross-track redundancy (NOT across patients) — see note above",
        "Variant classification: HGVS-derived; insertion / deletion / duplication / delins / repeat are distinct classes",
        "Patient metadata: parsed from pub.html (Sex, Age, Disease, Symptoms, References)",
        "Gender normalized: XY→M, XX→F; ambiguous values flagged in _gender_note",
        "Pending: audit of premature-stop-codon variants where Mutalyzer RNA annotation may be unreliable",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value="• " + note)
        r += 1

    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 12


def main():
    df = pd.read_csv(IN_PATH, sep="\t", dtype=str).fillna("")
    df["_gender_note"] = df["Individual/Gender"].apply(classify_gender)

    variants = (
        df.drop_duplicates(["_gene", "_dedup_key"])
          .reindex(columns=VARIANT_COLS)
          .reset_index(drop=True)
    )

    indiv_groups = df.groupby("Individual/ID", sort=False)
    indiv = indiv_groups.agg({
        "_gene":                        agg_unique_join,
        "_accession":                   agg_unique_join,
        "Individual/Gender":            agg_unique_join,
        "Individual/Age":               agg_unique_join,
        "Individual/Origin/Geographic": agg_unique_join,
        "Individual/Remarks":           agg_unique_join,
        "Phenotype/Disease":            agg_unique_join,
        "Phenotype/Additional":         agg_unique_join,
        "Reference/PubMed_ID":          agg_unique_join,
        "Reference/Authors":            agg_unique_join,
        "Reference/Title":              agg_unique_join,
        "Reference/Location":           agg_unique_join,
        "_dedup_key":                   "size",
    }).rename(columns={"_dedup_key": "_allele_count"}).reset_index()
    indiv["_gender_note"] = indiv["Individual/Gender"].apply(classify_gender)
    indiv = indiv.reindex(columns=INDIVIDUAL_COLS)

    fill_rate_cols = [
        "Individual/Gender", "Individual/Age", "Individual/Origin/Geographic",
        "Individual/Remarks", "Phenotype/Disease", "Phenotype/Additional",
        "VariantOnGenome/DNA", "VariantOnGenome/Remarks",
        "transcriptid", "VariantOnTranscript/DNA", "VariantOnTranscript/Protein",
        "Reference/PubMed_ID", "Reference/Authors", "Reference/Title", "Reference/Location",
    ]
    fill_rates = {c: ((df[c] != "").sum(), len(df)) for c in fill_rate_cols}

    track_counts = variants["_source_track"].value_counts().to_dict()
    class_counts = variants["_variant_class"].value_counts().to_dict()

    excluded_oos, excluded_protein, excluded_nofasta = compute_excluded_genes(df)

    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("Summary")
    build_summary_sheet(ws_summary, df, variants, indiv, fill_rates,
                        track_counts, class_counts,
                        excluded_oos, excluded_protein, excluded_nofasta)

    ws_var = wb.create_sheet("Variants")
    write_df(ws_var, variants)

    ws_ind = wb.create_sheet("Individuals")
    write_df(ws_ind, indiv)

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"  Summary    : counts + variant classes + gene exclusions + fill rates + provenance")
    print(f"  Variants   : {len(variants):>5} rows  (one per unique variant, with _variant_class)")
    print(f"  Individuals: {len(indiv):>5} rows  (one per patient)")
    print()
    print(f"  Excluded OOS      : {len(excluded_oos):>2}  ({', '.join(excluded_oos)})")
    print(f"  Excluded protein  : {len(excluded_protein):>2}  ({', '.join(excluded_protein)})")
    print(f"  Excluded no_fasta : {len(excluded_nofasta):>2}  ({', '.join(excluded_nofasta)})")
    print()
    print(f"  Variant classes   : {dict(class_counts)}")


if __name__ == "__main__":
    main()
